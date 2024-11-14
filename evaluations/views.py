from django.shortcuts import render, get_object_or_404, redirect

# Create your views here.

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import HttpResponse, HttpResponseForbidden
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.views import View
from .models import Department, JobRank
from .forms import EmployeeForm, DepartmentForm, JobRankForm

from .models import Employee, Evaluation, Answer, Question, Department, Profile
from .forms import EvaluationForm

import openpyxl


@login_required
def home(request):
    user = request.user
    if user.is_superuser:
        return redirect('/admin/')
    
    if is_employee(user):
        return redirect('view_own_evaluations')
    elif is_general_manager(user):
        return redirect('employee_list')
    elif is_manager(user) :
        return redirect('employee_list')
    else:
        return render(request, 'evaluations/unauthorized.html')


JOB_RANK_GROUPS = {
    0: 'مدیر ارشد',  # Senior Manager
    1: 'مدیر',       # Manager
    2: 'کارمند',     # Employee
}

def is_employee(user):
    return user.groups.filter(name=JOB_RANK_GROUPS[2]).exists()

def is_manager(user):
    return user.groups.filter(name=JOB_RANK_GROUPS[1]).exists()

def is_general_manager(user):
    return user.groups.filter(name=JOB_RANK_GROUPS[0]).exists()



def user_department(user):
    return user.profile.department if hasattr(user, 'profile') else None


@login_required
def employee_list(request):
    if is_general_manager(request.user):
        employees = Employee.objects.all()
    elif is_manager(request.user):
        department = user_department(request.user)
        employees = Employee.objects.filter(department=department)
    else:
        raise PermissionDenied
    
    return render(request, 'evaluations/employee_list.html', {'employees': employees})


@login_required
def create_employee(request):
    if not is_general_manager(request.user):
        raise PermissionDenied
    
    if request.method == 'POST':
        file_number = request.POST.get('file_number')
        name = request.POST.get('name')
        job_title = request.POST.get('job_title')
        job_rank_id = request.POST.get('job_rank')
        job_rank = JobRank.objects.get(id=job_rank_id)
        department_id = request.POST.get('department')
        
        username = f"user_{file_number}"
        password = User.objects.make_random_password()
        if User.objects.filter(username=username).exists():
            # Handle the error (e.g., return an error message)
            return render(request, 'evaluations/create_employee.html', {
                'departments': departments,
                'error': 'نام کاربری قبلاً وجود دارد.'
            })
        user = User.objects.create_user(username=username, password=password)
        
        #employee_group = Group.objects.get(name='کارمند')
        #user.groups.add(employee_group)
        
        employee = Employee.objects.create(
            user=user,
            file_number=file_number,
            name=name,
            job_title=job_title,
            job_rank=job_rank,
            department_id=department_id
        )
        
        #if hasattr(user, 'profile'):
        #    user.profile.department_id = department_id
        #    user.profile.job_rank = job_rank
        #    user.profile.save()
        #else:
        #    Profile.objects.create(user=user, department_id=department_id)
        #    Profile.objects.create(user=user, job_rank=job_rank)
        if hasattr(user, 'profile'):
            user.profile.department = Department.objects.get(id=department_id)
            user.profile.job_rank = job_rank
            user.profile.save()

        
        # send_mail(
        #     'اطلاعات ورود به سیستم',
        #     f'نام کاربری: {username}\nرمز عبور: {password}',
        #     'from@example.com',
        #     [user.email],
        # )
        return redirect('employee_list')
    
    else:
        departments = Department.objects.all()
        return render(request, 'evaluations/create_employee.html', {'departments': departments})


@login_required
def evaluate_employee(request, employee_id):
    
    employee = get_object_or_404(Employee, id=employee_id)
    user = request.user
    
    if is_general_manager(user):
        pass
    elif is_manager(user):
        if employee.department != user_department(user):
            raise PermissionDenied
    else:
        raise PermissionDenied
    
    if request.method == 'POST':
        form = EvaluationForm(request.POST, job_rank=employee.job_rank)
        if form.is_valid():
            try:
                month_str = request.POST.get('month')
                if month_str and month_str.isdigit():
                    month = int(month_str)
                else:
                    month = timezone.now().month

                if not 1 <= month <= 12:
                    month = timezone.now().month
            except:
                month = 11
            evaluation = Evaluation.objects.create(
                employee=employee,
                evaluator=user,
                month=month
            )
            total_questions = 0
            scores = [0]*5  # Index 0 for 'عالی', Index 4 for 'ضعیف'
            for category in form.categories:
                for question in category['questions']:
                    field_name = f"question_{question.id}"
                    value = form.cleaned_data.get(field_name)
                    choice = int(value)
                    Answer.objects.create(
                        evaluation=evaluation,
                        question=question,
                        choice=choice
                    )
                    index = 5 - choice  # معکوس
                    scores[index] += 1
                    total_questions += 1
            # محاسبه امتیاز
            weights = [100, 80, 70, 60, 50]
            if total_questions == 0:
                total_score = 0
            else:
                
                weighted_scores = [(scores[i] * weights[i]) / total_questions for i in range(5)]
                total_score = sum(weighted_scores)
            
            evaluation.scores = {
                'individual_scores': weighted_scores,
                'total_score': total_score
            }
            evaluation.total_score = total_score
            evaluation.save()
            return redirect('view_evaluation', evaluation_id=evaluation.id)
    else:
        form = EvaluationForm(job_rank=employee.job_rank)
    
    return render(request, 'evaluations/evaluate_employee.html', {
        'employee': employee,
        'form': form,
        'categories': form.categories, 
    })


@login_required
def view_evaluation(request, evaluation_id):
    evaluation = get_object_or_404(Evaluation, id=evaluation_id)
    user = request.user
    
    if is_general_manager(user):
        pass
    elif is_manager(user):
        if evaluation.employee.department != user_department(user):
            raise PermissionDenied
    elif is_employee(user):
        if evaluation.employee.user != user:
            raise PermissionDenied
    else:
        raise PermissionDenied
    
    answers = Answer.objects.filter(evaluation=evaluation)
    return render(request, 'evaluations/view_evaluation.html', {
        'evaluation': evaluation,
        'answers': answers,
    })


@login_required
def export_evaluation_to_excel(request, evaluation_id):
    evaluation = get_object_or_404(Evaluation, id=evaluation_id)
    user = request.user
    
    if is_general_manager(user):
        pass
    
    elif is_manager(user):
        if evaluation.employee.department != user_department(user):
            raise PermissionDenied
    
    elif is_employee(user):
        if evaluation.employee.user != user:
            raise PermissionDenied
    else:
        raise PermissionDenied
    
    answers = Answer.objects.filter(evaluation=evaluation)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Evaluation'
    
    sheet['A1'] = 'شرح عوامل ارزیابی'
    sheet['B1'] = 'امتیاز'
    
    idx = 2
    for answer in answers:
        sheet.cell(row=idx, column=1).value = answer.question.text
        sheet.cell(row=idx, column=2).value = answer.get_choice_display()
        idx += 1
    
    sheet.cell(row=idx, column=1).value = 'مجموع امتیازات'
    sheet.cell(row=idx, column=2).value = evaluation.total_score
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"evaluation_{evaluation.employee.file_number}_{evaluation.month}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def view_own_evaluations(request):
    user = request.user
    if not is_employee(user):
        raise PermissionDenied
    
    evaluations = Evaluation.objects.filter(employee__user=user)
    return render(request, 'evaluations/own_evaluations.html', {'evaluations': evaluations})

@login_required
def create_department(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create_employee')
    else:
        form = DepartmentForm()
    return render(request, 'evaluations/create_department.html', {'form': form})

@login_required
def create_job_rank(request):
    if request.method == 'POST':
        form = JobRankForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create_employee')
    else:
        form = JobRankForm()
    return render(request, 'evaluations/create_job_rank.html', {'form': form})

#cloner174